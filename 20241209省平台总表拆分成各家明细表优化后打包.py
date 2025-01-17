import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import tkinter as tk
from tkinter import filedialog, messagebox

# 使用 tkinter 进行确认框和文件选择
root = tk.Tk()
root.withdraw()  # 隐藏主窗口

# 弹出确认框，询问用户是否继续执行
confirm = messagebox.askyesno("确认", "是否继续执行文件选择操作？")

if confirm:
    # 选择输入文件（Excel 文件）
    input_file = filedialog.askopenfilename(title="选择一个Excel文件", filetypes=[("Excel Files", "*.xlsx")])

    # 选择输出目录
    output_dir = filedialog.askdirectory(title="选择输出文件夹")

    # 如果没有选择文件或文件夹，提前退出
    if not input_file or not output_dir:
        messagebox.showerror("错误", "必须选择文件和输出目录！")
        exit()

    # 定义列宽
    column_widths = {
        '医疗机构编码': 6.25,
        '医疗机构名称': 10,
        '患者姓名': 4,
        '患者性别': 5.8,
        '险种类型': 8,
        '结算日期': 6.6,
        '医保目录名称': 4,
        '规则名称': 9,
        '疑似违规内容': 13.25,
        '疑似违规金额': 6,
        '初审意见': 15,
        '复审意见': 15,
        '终审意见': 15,
        '申诉意见': 6.8,
        '终审结论': 3.4,
        '扣款金额（元）': 5.7,
        '终审时间': 10,
        '二次反馈': 4,
        '备注': 6,
    }

    # 定义需要保留的列及其顺序
    columns_to_keep = [
        '医疗机构编码', '医疗机构名称', '患者姓名', '患者性别', '险种类型',
        '结算日期', '医保目录名称', '规则名称', '疑似违规内容',
        '疑似违规金额', '初审意见', '申诉意见', '复审意见',
        '终审结论', '终审意见', '扣款金额（元）', '终审时间',
        '二次反馈', '备注'
    ]

    # 从选中的文件读取数据
    file_name = os.path.basename(input_file)

    # 从文件名中提取年和月份信息
    date_match = re.search(r'(\d{4})(\d{2})', file_name)
    if date_match:
        year = date_match.group(1)
        month = date_match.group(2).lstrip('0')  # 去掉月份的前导零
        main_title = f"淮安经济技术开发区智能审核{year}{month}扣款明细统计表"
    else:
        main_title = "淮安经济技术开发区智能审核扣款明细统计表"  # 默认标题

    # 读取数据，第二行是标题行
    df = pd.read_excel(input_file, header=1)  # 修改为header=1
    df.dropna(how='all', inplace=True)

    # 清理列名
    df.columns = df.columns.str.strip()

    # 检查“扣款金额”列
    if '扣款金额' in df.columns:
        df = df[df['扣款金额'] != 0]
        df['扣款金额'] = pd.to_numeric(df['扣款金额'], errors='coerce').round(2)

        # 根据医疗机构编码分组
        for medical_institution_code in df['医疗机构编码'].unique():
            institution_data = df[df['医疗机构编码'] == medical_institution_code].copy()

            # 删除“医保目录编码”这一列
            institution_data.drop(columns=['医保目录编码'], inplace=True, errors='ignore')

            # 创建“扣款金额（元）”列
            institution_data.rename(columns={'扣款金额': '扣款金额（元）'}, inplace=True)

            # 添加“二次反馈”列并填入“无异议”
            institution_data['二次反馈'] = '无异议'

            # 添加“备注”列，内容为空
            institution_data['备注'] = ''

            # 根据需要保留的列进行筛选
            institution_data = institution_data[columns_to_keep]

            # 使用第二个医疗机构名称来命名文件
            institution_names = institution_data['医疗机构名称'].unique()
            institution_name = institution_names[1] if len(institution_names) > 1 else institution_names[0]
            output_file = os.path.join(output_dir, f'{institution_name}.xlsx')

            # 保存数据到 Excel
            institution_data.to_excel(output_file, index=False)

            # 添加总标题
            workbook = load_workbook(output_file)
            sheet = workbook.active

            # 页面设置
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.fitToPage = True
            sheet.page_setup.fitToHeight = 1
            sheet.page_setup.fitToWidth = 1

            sheet.insert_rows(0)
            sheet['A1'] = main_title
            sheet['A1'].font = Font(name='方正小标宋_GBK', bold=True, size=24)
            sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
            sheet.row_dimensions[1].height = 65

            # 查找小标题的位置并合并单元格
            first_col = last_col = None
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=2, column=col).value
                if cell_value is not None:
                    if first_col is None:
                        first_col = col
                    last_col = col
            if first_col is not None and last_col is not None:
                merge_range = f"{sheet.cell(row=1, column=first_col).coordinate}:{sheet.cell(row=1, column=last_col).coordinate}"
                sheet.merge_cells(merge_range)

            # 设置小标题格式
            for col in range(1, sheet.max_column + 1):
                small_title_cell = sheet.cell(row=2, column=col)
                small_title_cell.font = Font(name='黑体', size=12)
                small_title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 设置列宽和样式
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            for col_name, width in column_widths.items():
                try:
                    col_index = institution_data.columns.get_loc(col_name) + 1
                    col_letter = get_column_letter(col_index)
                    sheet.column_dimensions[col_letter].width = width
                except KeyError:
                    print(f"列名 '{col_name}' 在数据中未找到，跳过该列。")

            # 计算扣款金额（元）列的总金额
            total_amount = institution_data['扣款金额（元）'].sum()

            # 获取扣款金额列所在的列号
            扣款金额列 = '扣款金额（元）'
            扣款金额列索引 = institution_data.columns.get_loc(扣款金额列)

            # 获取最后一行的行号
            last_row = len(institution_data) + 2  # 数据行数 + 2（因为标题行）

            # 在扣款金额左边插入 "违规总金额："
            sheet.cell(row=last_row + 1, column=扣款金额列索引, value="违规总金额：")

            # 在扣款金额的单元格插入总金额
            sheet.cell(row=last_row + 1, column=扣款金额列索引 + 1, value=total_amount)

            # 在“违规总金额：”下方插入“经办人签字：”和“盖章：”在同一行，且不设置边框
            sheet.cell(row=last_row + 3, column=扣款金额列索引 - 2, value="经办人签字：")
            sheet.cell(row=last_row + 3, column=扣款金额列索引 + 2, value="盖章：")

            # 去除“经办人签字”和“盖章”这一行的边框
            for row in sheet.iter_rows(min_row=last_row + 3, max_row=last_row + 3):
                for cell in row:
                    cell.border = Border()  # 去掉边框

            # 为除了主标题外的所有单元格添加边框，但跳过"违规总金额："和"经办人签字："这一行
            for row in sheet.iter_rows(min_row=1):
                for cell in row:
                    if cell.row == last_row + 1 or cell.row == last_row + 2 or cell.row == last_row + 3:
                        continue  # 跳过这一行，保持没有边框
                    if isinstance(cell.value, (int, float)):
                        cell.font = Font(name='Times New Roman')
                        cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # 为除了主标题外的所有单元格添加边框
                    if cell.row > 1:
                        cell.border = thin_border

            workbook.save(output_file)

        # 文件处理完成，弹出保存成功的提示框
        messagebox.showinfo("完成", f"文件已成功保存至：{output_dir}")

    else:
        print(f"文件 '{file_name}' 中未找到 '扣款金额' 列，跳过该文件。")

else:
    print("操作已取消。")
