import pandas as pd
import os
from openpyxl.styles import Alignment, Font, Border, Side
from tkinter import Tk, filedialog
import tkinter.messagebox as messagebox
import re

# 定义边框样式
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# 定义字体样式
body_font_chinese = Font(name='方正仿宋_GBK')  # 汉字字体
body_font_numbers = Font(name='Times New Roman')  # 数字字体

# 定义排序顺序
custom_order = ['H32087100006', 'H32087100010', 'H32087100021', 'H32087100196', 'H32087101766']

# 修改合并和分组函数
def merge_and_group(df_temp):
    # 替换医院名称，避免SettingWithCopyWarning
    df_temp.loc[:, '医疗机构名称'] = df_temp['医疗机构名称'].replace({
        '淮安经济技术开发区医院（淮安汉方医院管理有限公司）': '淮安经济技术开发区医院',
        '枚乘路社区卫生服务中心': '淮安经济技术开发区枚乘街道卫生院'
    })

    # 先按编码进行分组
    grouped = df_temp.groupby('医疗机构编码', as_index=False).agg({
        '医疗机构名称': lambda x: x.loc[x.str.len().idxmax()],  # 保留名称最长的医院名称
        '扣款金额': 'sum',
        '人次': 'sum'
    })

    # 按照给定的排序顺序重新排列
    ordered_rows = []
    remaining_rows = []

    # 分离出根据 custom_order 排序的行
    for _, row in grouped.iterrows():
        if row['医疗机构编码'] in custom_order:
            ordered_rows.append(row)
        else:
            remaining_rows.append(row)

    # 将ordered_rows和remaining_rows转回为DataFrame
    ordered_rows = pd.DataFrame(ordered_rows)
    remaining_rows = pd.DataFrame(remaining_rows)

    # 确保remaining_rows有'医疗机构编码'列，并按编码排序
    if '医疗机构编码' in remaining_rows.columns:
        remaining_rows = remaining_rows.sort_values(by='医疗机构编码')
    else:
        print("Warning: '医疗机构编码' column is missing in remaining rows.")

    # 合并排序后的数据
    grouped = pd.concat([ordered_rows, remaining_rows], ignore_index=True)

    return grouped

def main():
    root = Tk()
    root.withdraw()

    messagebox.showinfo("选择文件", "请选择需要处理的 Excel 文件")
    input_files = filedialog.askopenfilenames(title="选择 Excel 文件", filetypes=[("Excel 文件", "*.xlsx")])
    if not input_files:
        messagebox.showerror("错误", "未选择任何文件")
        return

    messagebox.showinfo("选择输出文件夹", "请选择输出文件保存的位置")
    output_path = filedialog.askdirectory(title="选择输出文件夹")
    if not output_path:
        messagebox.showerror("错误", "未选择输出文件夹")
        return

    for file in input_files:
        df = pd.read_excel(file, header=1)

        if '人次' not in df.columns:
            df['人次'] = 1
        needed_columns = ['医疗机构编码', '医疗机构名称', '人次', '扣款金额', '险种类型']
        df = df[needed_columns]

        df['扣款金额'] = pd.to_numeric(df['扣款金额'], errors='coerce')
        df['人次'] = pd.to_numeric(df['人次'], errors='coerce')

        file_name = os.path.basename(file).split('.')[0]
        year_month = file_name.split('智能审核')[1][:6]
        year = year_month[:4]
        month = year_month[4:]

        for insurance_type in df['险种类型'].unique():
            df_temp = df[df['险种类型'] == insurance_type]
            df_grouped = merge_and_group(df_temp)

            df_grouped.insert(0, '序号', range(1, len(df_grouped) + 1))

            total_row = pd.DataFrame({
                '序号': ['合计'],
                '医疗机构编码': [''],
                '医疗机构名称': [''],
                '人次': [df_grouped['人次'].sum()],
                '扣款金额': [df_grouped['扣款金额'].sum()]
            })
            df_grouped = pd.concat([df_grouped, total_row], ignore_index=True)

            output_file_name = f"淮安经济技术开发区智能审核{year}年{month}月扣款统计表{insurance_type}.xlsx"
            output_file_path = os.path.join(output_path, output_file_name)

            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                df_grouped.to_excel(writer, index=False, startrow=2)

                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                title = f"淮安经济技术开发区智能审核{year}年{month}月扣款统计表（{insurance_type}）"
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_grouped.columns))
                worksheet.cell(row=1, column=1).value = title
                worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
                worksheet.cell(row=1, column=1).font = Font(name='方正小标宋_GBK', size=22)
                worksheet.row_dimensions[1].height = 65

                worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df_grouped.columns))
                worksheet.cell(row=2, column=1).value = "单位：人次/元"
                worksheet.cell(row=2, column=1).alignment = Alignment(horizontal='right', vertical='center')
                worksheet.cell(row=2, column=1).font = body_font_chinese  # 设置"单位：人次/元"的字体为方正仿宋_GBK

                worksheet.column_dimensions['A'].width = 8.4
                worksheet.column_dimensions['B'].width = 20
                worksheet.column_dimensions['C'].width = 56
                worksheet.column_dimensions['D'].width = 12
                worksheet.column_dimensions['E'].width = 12
                worksheet.column_dimensions['F'].width = 12  # 增加了 F 列宽度

                header_font = Font(name='黑体', bold=True)
                for col in range(1, len(df_grouped.columns) + 1):
                    worksheet.cell(row=3, column=col).font = header_font

                # 设置正文字体（汉字为方正仿宋_GBK，数字为Times New Roman）
                for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=1,
                                               max_col=len(df_grouped.columns)):
                    for cell in row:
                        if cell.column == 2:  # 医疗机构编码列
                            if cell.row == 3:
                                cell.font = Font(name='黑体', bold=True)
                            else:
                                cell.font = body_font_numbers  # 编码数据使用 Times New Roman
                        else:
                            if isinstance(cell.value, (int, float)):
                                cell.font = body_font_numbers
                            else:
                                cell.font = body_font_chinese
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                last_row = worksheet.max_row
                worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=3)
                worksheet.cell(row=last_row, column=1).value = "合计"
                worksheet.cell(row=last_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

                # 修改审批人、复核、初审的列位置
                worksheet.cell(row=last_row + 1, column=2).value = "审批人："
                worksheet.cell(row=last_row + 1, column=3).value = "复核："
                worksheet.cell(row=last_row + 1, column=4).value = "初审："  # 修改为第四列

                # 设置字体为 方正仿宋_GBK
                worksheet.cell(row=last_row + 1, column=2).font = body_font_chinese
                worksheet.cell(row=last_row + 1, column=3).font = body_font_chinese
                worksheet.cell(row=last_row + 1, column=4).font = body_font_chinese  # 初审列也设置字体

                # 设置对齐方式
                for col in [2, 3, 4]:  # 更新为包括第4列
                    worksheet.cell(row=last_row + 1, column=col).alignment = Alignment(vertical='center')

    messagebox.showinfo("完成", "所有文件已处理并保存！")

if __name__ == "__main__":
    main()
